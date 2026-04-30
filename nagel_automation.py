#!/usr/bin/env python3
"""
nagel_automation.py — Nagel Law Expense Automation
===================================================
Runs nightly at 7:00 PM via GitHub Actions.

What it does:
  1. Scans each entity folder in Google Drive for new documents
  2. Sends each file to Claude AI which extracts: vendor, amount,
     date, category, entity, invoice number, and confidence score
  3. Confidence >= 90% → appended to Firm_Expense_Tracker.xlsx
  4. Confidence <  90% → saved to a "Needs Review" sheet for Andres
  5. Processed files moved to /done subfolder
  6. New entity folders detected automatically — no code changes needed
  7. Summary email sent to andres@nagellaw.com

Setup:
  pip install google-api-python-client google-auth openpyxl anthropic pillow

Environment variables (set in GitHub Actions secrets):
  GOOGLE_CREDENTIALS_JSON  — contents of your service account JSON key file
  ANTHROPIC_API_KEY        — your Anthropic API key
  SUMMARY_EMAIL            — andres@nagellaw.com
  SMTP_USER                — Gmail address used to send the summary
  SMTP_PASSWORD            — Gmail App Password (not your regular password)
  DRIVE_ROOT_FOLDER_ID     — ID of the "Nagel Law — Intake" Drive folder
  EXCEL_FILE_ID            — ID of Firm_Expense_Tracker.xlsx on Drive
"""

import os
import io
import json
import base64
import tempfile
import smtplib
import logging
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Google
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload

# Excel
import openpyxl

# Anthropic
import anthropic

# ─── Logging ────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
log = logging.getLogger("nagel")

# ─── Config ─────────────────────────────────────────────────────────────────
ENTITIES = ["AFLE", "GT Nevis", "GT Bank", "Nagel & Associates"]
UNRECOGNIZED_FOLDER = "_Unrecognized"
DONE_FOLDER = "done"
CONFIDENCE_THRESHOLD = 0.90
SUMMARY_EMAIL = os.environ.get("SUMMARY_EMAIL", "andres@nagellaw.com")
DRIVE_ROOT_ID = os.environ.get("DRIVE_ROOT_FOLDER_ID", "")
EXCEL_FILE_ID = os.environ.get("EXCEL_FILE_ID", "")

SUPPORTED_MIME = {
    "application/pdf",
    "image/jpeg",
    "image/png",
    "image/heic",
    "image/webp",
    "image/tiff",
}

CATEGORIES = [
    "Payroll",
    "Rent & Utilities",
    "Insurance",
    "Professional Services",
    "Marketing",
    "Travel & Meals",
    "Software & Subscriptions",
    "Office Supplies",
    "Bank & Merchant Fees",
    "Equipment",
    "Taxes & Licenses",
    "Other",
]

# ─── Google Drive client ─────────────────────────────────────────────────────
def get_drive_service():
    creds_json = os.environ.get("GOOGLE_CREDENTIALS_JSON", "")
    if not creds_json:
        raise ValueError("GOOGLE_CREDENTIALS_JSON env var not set.")
    info = json.loads(creds_json)
    creds = service_account.Credentials.from_service_account_info(
        info,
        scopes=["https://www.googleapis.com/auth/drive"]
    )
    return build("drive", "v3", credentials=creds)


# ─── Drive helpers ───────────────────────────────────────────────────────────
def list_subfolders(drive, parent_id):
    """Return {name: id} for all subfolders of parent_id. Supports Shared Drives."""
    result = drive.files().list(
        q=f"'{parent_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false",
        fields="files(id, name)",
        includeItemsFromAllDrives=True,
        supportsAllDrives=True,
        corpora="allDrives"
    ).execute()
    return {f["name"]: f["id"] for f in result.get("files", [])}


def list_files(drive, folder_id):
    """Return list of file dicts in a folder. Supports Shared Drives."""
    result = drive.files().list(
        q=f"'{folder_id}' in parents and mimeType!='application/vnd.google-apps.folder' and trashed=false",
        fields="files(id, name, mimeType, size)",
        includeItemsFromAllDrives=True,
        supportsAllDrives=True,
        corpora="allDrives"
    ).execute()
    return result.get("files", [])


def ensure_subfolder(drive, parent_id, name):
    """Get or create a subfolder by name. Supports Shared Drives."""
    existing = list_subfolders(drive, parent_id)
    if name in existing:
        return existing[name]
    meta = {
        "name": name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [parent_id]
    }
    folder = drive.files().create(
        body=meta,
        fields="id",
        supportsAllDrives=True
    ).execute()
    log.info(f"Created subfolder '{name}' in Drive.")
    return folder["id"]


def download_file(drive, file_id):
    """Download a Drive file into a BytesIO buffer. Supports Shared Drives."""
    buf = io.BytesIO()
    request = drive.files().get_media(
        fileId=file_id,
        supportsAllDrives=True
    )
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return buf


def find_file_in_folder(drive, folder_id, filename):
    """Search for a file by name inside a folder. Supports Shared Drives."""
    result = drive.files().list(
        q=f"'{folder_id}' in parents and name='{filename}' and trashed=false",
        fields="files(id, name)",
        includeItemsFromAllDrives=True,
        supportsAllDrives=True,
        corpora="allDrives"
    ).execute()
    files = result.get("files", [])
    return files[0]["id"] if files else None


def move_file(drive, file_id, new_parent_id, old_parent_id):
    """Move a file to a different folder. Supports Shared Drives."""
    drive.files().update(
        fileId=file_id,
        addParents=new_parent_id,
        removeParents=old_parent_id,
        fields="id, parents",
        supportsAllDrives=True
    ).execute()


def export_portal_json(wb, output_path="data.json"):
    """
    Read the Transactions and Needs Review sheets and write data.json
    to the local filesystem. GitHub Actions will git-push it to the repo
    so the portal reads live data automatically.
    """
    import json

    # ── Transactions ──
    ws = wb["Transactions"] if "Transactions" in wb.sheetnames else None
    if not ws:
        log.warning("No Transactions sheet found — skipping JSON export.")
        return

    transactions = []
    for row_num in range(3, ws.max_row + 1):
        date_val = ws.cell(row=row_num, column=1).value
        vendor_val = ws.cell(row=row_num, column=5).value
        if not date_val and not vendor_val:
            continue
        try:
            amount = float(ws.cell(row=row_num, column=8).value or 0)
        except:
            amount = 0
        transactions.append({
            "date":         str(ws.cell(row=row_num, column=1).value  or ""),
            "year":         str(ws.cell(row=row_num, column=2).value  or ""),
            "month":        str(ws.cell(row=row_num, column=3).value  or ""),
            "entity":       str(ws.cell(row=row_num, column=4).value  or ""),
            "vendor":       str(ws.cell(row=row_num, column=5).value  or ""),
            "category":     str(ws.cell(row=row_num, column=6).value  or ""),
            "desc":         str(ws.cell(row=row_num, column=7).value  or ""),
            "amount":       amount,
            "status":       str(ws.cell(row=row_num, column=9).value  or "Pending"),
            "source":       str(ws.cell(row=row_num, column=10).value or ""),
            "inv":          str(ws.cell(row=row_num, column=11).value or ""),
            "filename":     str(ws.cell(row=row_num, column=12).value or ""),
            "due_date":     str(ws.cell(row=row_num, column=13).value or ""),
            "payment_date": str(ws.cell(row=row_num, column=14).value or ""),
        })

    # ── Needs Review ──
    review = []
    if "Needs Review" in wb.sheetnames:
        ws_r = wb["Needs Review"]
        for row_num in range(3, ws_r.max_row + 1):
            entity_val = ws_r.cell(row=row_num, column=2).value
            if not entity_val:
                continue
            try:
                amt = float(ws_r.cell(row=row_num, column=5).value or 0)
            except:
                amt = 0
            review.append({
                "date_processed": str(ws_r.cell(row=row_num, column=1).value or ""),
                "entity":         str(ws_r.cell(row=row_num, column=2).value or ""),
                "filename":       str(ws_r.cell(row=row_num, column=3).value or ""),
                "vendor":         str(ws_r.cell(row=row_num, column=4).value or ""),
                "amount":         amt,
                "date":           str(ws_r.cell(row=row_num, column=6).value or ""),
                "category":       str(ws_r.cell(row=row_num, column=7).value or ""),
                "inv":            str(ws_r.cell(row=row_num, column=8).value or ""),
                "desc":           str(ws_r.cell(row=row_num, column=9).value or ""),
                "confidence":     str(ws_r.cell(row=row_num, column=10).value or ""),
                "notes":          str(ws_r.cell(row=row_num, column=11).value or ""),
                "status":         str(ws_r.cell(row=row_num, column=12).value or ""),
            })

    payload = {
        "generated": datetime.today().strftime("%Y-%m-%d %H:%M"),
        "transactions": transactions,
        "needs_review": review
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    log.info(f"data.json written — {len(transactions)} transactions, {len(review)} needs review.")


def upload_excel(drive, local_path):
    """
    Upload updated Excel back to Drive.
    Searches for the file by name inside the root intake folder.
    If found — updates it in place.
    If not found — creates a new one inside the root folder.
    """
    excel_name = "Firm_Expense_Tracker.xlsx"
    media = MediaFileUpload(
        local_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False
    )

    # Search for the file by name inside the intake folder
    found_id = find_file_in_folder(drive, DRIVE_ROOT_ID, excel_name)

    if found_id:
        log.info(f"Found Excel in Drive (id={found_id}), updating...")
        drive.files().update(
            fileId=found_id,
            media_body=media,
            supportsAllDrives=True
        ).execute()
        log.info("Excel file updated on Drive successfully.")
    else:
        log.info("Excel not found in intake folder — creating new file...")
        file_meta = {
            "name": excel_name,
            "parents": [DRIVE_ROOT_ID]
        }
        drive.files().create(
            body=file_meta,
            media_body=media,
            fields="id",
            supportsAllDrives=True
        ).execute()
        log.info("New Excel file created in intake folder.")


# ─── AI extraction ───────────────────────────────────────────────────────────
def load_supplier_profile(entity: str, vendor_hint: str = "") -> dict:
    """
    Load a supplier profile from suppliers.json (versioned format).
    Matches by entity + vendor_name or any alias.
    Returns the profile dict or {} if not found.
    """
    import json as _json
    if not os.path.exists("suppliers.json"):
        return {}
    try:
        with open("suppliers.json") as f:
            db = _json.load(f)
        suppliers = db.get("suppliers", [])
        vh = (vendor_hint or "").strip().lower()
        for s in suppliers:
            if s.get("entity", "") != entity:
                continue
            names = [s.get("vendor_name", "")] + s.get("aliases", [])
            for name in names:
                nl = name.strip().lower()
                if not nl:
                    continue
                if nl == vh or (vh and (nl in vh or vh in nl)):
                    return s
        # No vendor match — return first supplier for this entity if any
        for s in suppliers:
            if s.get("entity", "") == entity:
                return s
        return {}
    except Exception as e:
        log.warning(f"Could not load supplier profile: {e}")
        return {}


def build_supplier_hint(profile: dict) -> str:
    """Convert a versioned supplier profile into a hint string for the AI prompt."""
    if not profile:
        return ""
    lines = ["\n\n=== SUPPLIER PROFILE ==="]
    lines.append(f"Vendor: {profile.get('vendor_name','')}")

    ctx = profile.get("vendor_context", "")
    if ctx:
        lines.append(f"Context: {ctx}")

    struct = profile.get("invoice_structure", {})
    if struct.get("vat_included"):
        rate = int(struct.get("vat_rate", 0) * 100)
        lines.append(f"VAT: {rate}% included. Distribute proportionally across category groups so groups sum to Total USD.")
    if struct.get("sends_paid_receipts"):
        lines.append(f"Paid receipts: {struct.get('paid_receipt_format','')}")
    if struct.get("invoice_number_location"):
        lines.append(f"Invoice # location: {struct.get('invoice_number_location')}")

    inv_types = profile.get("invoice_types", [])
    if inv_types:
        lines.append("Invoice types this vendor sends:")
        for t in inv_types:
            kws = ", ".join(t.get("keywords", []))
            cat = t.get("category", "MIXED")
            lines.append(f"  - {t.get('type_id','')} ({cat}): keywords [{kws}]")

    rules = profile.get("line_item_categorization_rules", [])
    if rules:
        lines.append("Line-item categorization rules (match keywords against each line item):")
        for r in rules:
            kws = ", ".join(r.get("match_keywords", []))
            lines.append(f"  - {r.get('category','')}: [{kws}]")

    extraction = profile.get("extraction_rules", {})
    approach = extraction.get("approach", "")
    if approach == "smart_grouping":
        lines.append("APPROACH: smart_grouping — create ONE entry per master category. "
                     "Sum line items within each category, then distribute VAT proportionally.")
        if extraction.get("category_default_warning"):
            lines.append(f"WARNING: {extraction['category_default_warning']}")

    hints = profile.get("confidence_hints", {})
    if isinstance(hints, dict) and hints:
        lines.append("Confidence hints:")
        for k, v in hints.items():
            lines.append(f"  - {k}: {v}")
    return "\n".join(lines)


def extract_invoice_data(file_bytes: bytes, filename: str, entity: str, mime_type: str) -> dict:
    """
    Send the document to Claude and get structured invoice data back.
    Returns a dict with keys: vendor, amount, date, category, invoice_number,
    description, confidence, notes
    """
    client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY", ""))

    # Convert to base64
    b64 = base64.standard_b64encode(file_bytes).decode("utf-8")

    # Normalise mime for Claude (HEIC not supported — convert hint)
    claude_mime = mime_type
    if mime_type not in {"image/jpeg", "image/png", "image/gif", "image/webp",
                         "application/pdf"}:
        claude_mime = "image/jpeg"   # fallback — PIL conversion happens below

    categories_list = "\n".join(f"- {c}" for c in CATEGORIES)

    # Load supplier profile for this entity (alias match happens inside)
    supplier_profile = load_supplier_profile(entity)
    supplier_hint = build_supplier_hint(supplier_profile)

    prompt = f"""You are an expert bookkeeper reviewing a business expense document for Nagel Law.

The document was uploaded to the "{entity}" entity folder.{supplier_hint}

CATEGORY SELECTION RULES — read carefully before choosing:
- "Rent & Utilities": office space rent, MRLO office provision, utility bills
- "Professional Services": legal advice, attorney fees, consulting, retainer fees for services (NOT office space)
- "Taxes & Licenses": government fees, incorporation fees, licence fees (ALHL, TCSP, Class II), stamp duty, filing fees, due diligence fees paid to government
- "Payroll": salaries, wages, payroll processing
- "Insurance": malpractice insurance, liability insurance, any insurance premium
- "Software & Subscriptions": software licences, SaaS subscriptions, online tools
- "Office Supplies": stationery, printer supplies, office consumables
- "Travel & Meals": flights, hotels, meals, transportation
- "Equipment": computers, hardware, furniture, machinery
- "Bank & Merchant Fees": wire transfer fees, bank charges, credit card processing fees
- "Marketing": advertising, campaigns, branding
- "Other": anything that does not fit the above

IMPORTANT RULES:
1. If an invoice has MULTIPLE line items from different categories, choose the category
   representing the LARGEST dollar amount on the invoice.
2. Government application fees, stamp duty, licence fees = "Taxes & Licenses" always.
3. Monthly office/MRLO retainer = "Rent & Utilities" not Professional Services.
4. Legal advice and attorney consultation = "Professional Services".

Extract the following:
1. vendor — the company or person being paid
2. amount — the TOTAL amount in USD including all taxes and fees (number only, no $ sign)
3. date — invoice or transaction date in YYYY-MM-DD format
4. category — pick the single best match using the rules above
5. invoice_number — the invoice or reference number if present, else "N/A"
6. description — a short 5-10 word description of the PRIMARY service or expense
7. due_date — payment due date in YYYY-MM-DD format, or null if not found
8. is_paid — true if Balance Due = $0.00 or shows a negative payment line or PAID stamp
9. payment_date — date payment was made in YYYY-MM-DD, or null
10. confidence — 0.0 to 1.0. Be strict: 0.95+ only if every field is clear
11. notes — if confidence below 0.90, briefly explain what is unclear

MULTI-CATEGORY GROUPING:
If the invoice has line items from different categories, create one entry per category group.
Each group should sum to its portion of the total (including proportional VAT).
The groups must sum to the invoice Total USD.

Respond ONLY with a valid JSON object, no markdown, no explanation:
{{
  "vendor": "...",
  "amount": 0.00,
  "date": "YYYY-MM-DD",
  "invoice_number": "...",
  "is_paid": false,
  "payment_date": "YYYY-MM-DD or null",
  "due_date": "YYYY-MM-DD or null",
  "confidence": 0.00,
  "notes": "...",
  "line_groups": [
    {{
      "category": "...",
      "amount": 0.00,
      "description": "5-10 word description of this group"
    }}
  ]
}}"""

    try:
        if claude_mime == "application/pdf":
            content = [
                {
                    "type": "document",
                    "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": b64
                    }
                },
                {"type": "text", "text": prompt}
            ]
        else:
            content = [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": claude_mime,
                        "data": b64
                    }
                },
                {"type": "text", "text": prompt}
            ]

        response = client.messages.create(
            model="claude-opus-4-6",
            max_tokens=1500,
            messages=[{"role": "user", "content": content}]
        )

        raw = response.content[0].text.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        data = json.loads(raw.strip())

        for key in ["vendor", "invoice_number", "confidence"]:
            if key not in data:
                raise ValueError(f"Missing field: {key}")

        data["confidence"] = float(data.get("confidence", 0))

        # Normalise line_groups — ensure they exist and amounts are floats
        groups = data.get("line_groups", [])
        if not groups:
            # Fallback: create single group from top-level fields
            groups = [{
                "category": data.get("category", "Other"),
                "amount":   float(data.get("amount", 0)),
                "description": data.get("description", "")
            }]
        else:
            for g in groups:
                g["amount"] = float(g.get("amount", 0))

        data["line_groups"] = groups
        data["amount"] = sum(g["amount"] for g in groups)
        return data

    except Exception as e:
        log.warning(f"AI extraction failed for {filename}: {e}")
        return {
            "vendor": "Unknown",
            "amount": 0.0,
            "date": datetime.today().strftime("%Y-%m-%d"),
            "category": "Other",
            "invoice_number": "N/A",
            "description": "Extraction failed",
            "confidence": 0.0,
            "notes": str(e)
        }


# ─── Excel helpers ───────────────────────────────────────────────────────────
TRANSACTIONS_HEADERS = [
    "Date", "Year", "Month", "Entity", "Vendor", "Category",
    "Description", "Amount (USD)", "Status", "Source",
    "Invoice #", "File Name", "Due Date", "Payment Date"
]

REVIEW_HEADERS = [
    "Date Processed", "Entity", "File Name",
    "Vendor (AI)", "Amount (AI)", "Date (AI)",
    "Category (AI)", "Invoice # (AI)", "Description (AI)",
    "Confidence", "AI Notes", "Status"
]


def get_or_create_sheet(wb, name, headers):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    ws.append(headers)
    return ws


def update_status_to_paid(ws, row_num, payment_date=None):
    """
    When a paid receipt is detected, update Status (col 9) to Paid
    and record the payment date in col 14.
    """
    current = ws.cell(row=row_num, column=9).value
    if str(current).strip().lower() != "paid":
        ws.cell(row=row_num, column=9).value = "Paid"
        if payment_date and str(payment_date) not in ("null", "", "None"):
            ws.cell(row=row_num, column=14).value = str(payment_date)
        log.info(f"  → Row {row_num} status updated to Paid"
                 + (f" | payment date: {payment_date}" if payment_date else ""))
    else:
        log.info(f"  → Row {row_num} already marked Paid — no change needed")


def first_empty_row(ws, start_row=3):
    """
    Find the first truly empty row starting from start_row.
    Skips title rows and header rows at the top.
    Default start_row=3 matches the Excel structure:
      Row 1 = Title banner
      Row 2 = Column headers
      Row 3+ = Data
    """
    for row_num in range(start_row, ws.max_row + 2):
        row_vals = [ws.cell(row=row_num, column=c).value for c in range(1, 12)]
        if all(v is None or str(v).strip() == "" for v in row_vals):
            return row_num
    return ws.max_row + 1


def is_duplicate(ws, filename, data=None, start_row=4):
    """
    Check for duplicates using TWO methods:

    Method 1 — Filename match:
      If the exact filename already exists in column 10, it's a duplicate.

    Method 2 — Content match (vendor + amount + date):
      Even if the filename is different, if vendor + amount + date all
      match an existing row it's the same invoice uploaded twice.
      Columns: 1=Date, 3=Vendor, 6=Amount

    Returns (True, reason) if duplicate, (False, "") if new.
    """
    for row_num in range(start_row, ws.max_row + 1):
        # Method 1: filename — column 12
        file_cell = ws.cell(row=row_num, column=12).value
        if file_cell and str(file_cell).strip() == str(filename).strip():
            return True, f"Filename already exists in row {row_num}", row_num

        # Method 2: vendor + invoice# + category
        # With multi-category grouping, same invoice can have multiple rows.
        # Duplicate = same vendor + invoice# + category.
        # Same invoice + different category = valid new row.
        if data:
            ex_vendor   = str(ws.cell(row=row_num, column=5).value or "").strip().lower()
            ex_inv      = str(ws.cell(row=row_num, column=11).value or "").strip()
            ex_category = str(ws.cell(row=row_num, column=6).value or "").strip()

            new_vendor   = str(data.get("vendor", "")).strip().lower()
            new_inv      = str(data.get("invoice_number", "")).strip()
            new_category = str(data.get("category", "")).strip()

            if not ex_inv or not new_inv or ex_inv in ("N/A","") or new_inv in ("N/A",""):
                continue

            if ex_vendor and new_vendor:
                vendor_match   = (ex_vendor == new_vendor or
                    (len(ex_vendor)>4 and len(new_vendor)>4 and
                     (ex_vendor in new_vendor or new_vendor in ex_vendor)))
                inv_match      = ex_inv == new_inv
                category_match = ex_category.lower() == new_category.lower()

                if vendor_match and inv_match and category_match:
                    return True, (
                        f"Duplicate in row {row_num} — "
                        f"vendor: '{ex_vendor}' / invoice#: {ex_inv} / category: {ex_category}"
                    ), row_num

    return False, "", None


def append_transaction(ws, entity, data, filename):
    """
    Append a confirmed row to the first empty row in Transactions sheet.
    Column layout (matches Firm_Expense_Tracker.xlsx):
      A(1)=Date  B(2)=Year  C(3)=Month  D(4)=Entity  E(5)=Vendor
      F(6)=Category  G(7)=Description  H(8)=Amount(USD)
      I(9)=Status  J(10)=Source  K(11)=Invoice#  L(12)=Filename
    """
    row = first_empty_row(ws)
    date_str = data.get("date", "")
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        year  = dt.year
        month = dt.strftime("%B")   # e.g. "January"
    except Exception:
        year  = ""
        month = ""

    due_date = data.get("due_date", None)
    if due_date == "null" or due_date == "":
        due_date = None

    ws.cell(row=row, column=1).value  = date_str
    ws.cell(row=row, column=2).value  = year
    ws.cell(row=row, column=3).value  = month
    ws.cell(row=row, column=4).value  = entity
    ws.cell(row=row, column=5).value  = data.get("vendor", "")
    ws.cell(row=row, column=6).value  = data.get("category", "")
    ws.cell(row=row, column=7).value  = data.get("description", "")
    ws.cell(row=row, column=8).value  = data.get("amount", 0)
    ws.cell(row=row, column=9).value  = "Pending"
    ws.cell(row=row, column=10).value = "Drive"
    ws.cell(row=row, column=11).value = data.get("invoice_number", "N/A")
    ws.cell(row=row, column=12).value = filename
    ws.cell(row=row, column=13).value = due_date
    log.info(f"  Written to row {row} — {entity} | {data.get('vendor')} | ${data.get('amount')} | due: {due_date}")


def append_review(ws, entity, data, filename):
    """Append a low-confidence row to the first empty row in Needs Review sheet."""
    row = first_empty_row(ws, start_row=3)
    ws.cell(row=row, column=1).value  = datetime.today().strftime("%Y-%m-%d")
    ws.cell(row=row, column=2).value  = entity
    ws.cell(row=row, column=3).value  = filename
    ws.cell(row=row, column=4).value  = data.get("vendor", "")
    ws.cell(row=row, column=5).value  = data.get("amount", "")
    ws.cell(row=row, column=6).value  = data.get("date", "")
    ws.cell(row=row, column=7).value  = data.get("category", "")
    ws.cell(row=row, column=8).value  = data.get("invoice_number", "")
    ws.cell(row=row, column=9).value  = data.get("description", "")
    ws.cell(row=row, column=10).value = f"{data.get('confidence', 0)*100:.0f}%"
    ws.cell(row=row, column=11).value = data.get("notes", "")
    ws.cell(row=row, column=12).value = "Needs Review"
    log.info(f"  Written to review row {row}")


def ensure_entity_in_excel(ws_entities, entity_name):
    """Add entity row to Entities sheet if not already there."""
    existing = [ws_entities.cell(row=r, column=1).value
                for r in range(2, ws_entities.max_row + 1)]
    if entity_name not in existing:
        # Generate a simple code from the name
        code = entity_name.upper().replace(" ", "-")[:8]
        ws_entities.append([entity_name, code])
        log.info(f"New entity added to Excel: {entity_name}")


# ─── Main pipeline ───────────────────────────────────────────────────────────
def run():
    log.info("=" * 60)
    log.info("Nagel Law — Expense Automation starting")
    log.info("=" * 60)

    drive = get_drive_service()
    ai_client_check = os.environ.get("ANTHROPIC_API_KEY", "")
    if not ai_client_check:
        raise ValueError("ANTHROPIC_API_KEY env var not set.")

    # ── Verify root Shared Drive access ──
    try:
        root_info = drive.files().get(
            fileId=DRIVE_ROOT_ID,
            fields="id, name",
            supportsAllDrives=True
        ).execute()
        log.info(f"Connected to Shared Drive: '{root_info.get('name', 'unknown')}'")
    except Exception as e:
        log.warning(f"Could not get root info: {e}")

    # ── Download current Excel from Drive ──
    log.info("Downloading Excel from Drive...")
    excel_name = "Firm_Expense_Tracker.xlsx"
    excel_id = find_file_in_folder(drive, DRIVE_ROOT_ID, excel_name)
    if not excel_id:
        raise FileNotFoundError(
            f"Could not find '{excel_name}' inside the Intake folder. "
            f"Make sure the file is inside the shared 'Nagel Law — Intake' folder "
            f"and the service account has Editor access to that folder."
        )
    log.info(f"Found Excel file (id={excel_id}), downloading...")
    excel_bytes = download_file(drive, excel_id)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(excel_bytes.read())
        excel_path = tmp.name

    wb = openpyxl.load_workbook(excel_path)

    # Ensure required sheets exist
    ws_tx     = get_or_create_sheet(wb, "Transactions", TRANSACTIONS_HEADERS)
    ws_review = get_or_create_sheet(wb, "Needs Review", REVIEW_HEADERS)
    ws_ents   = get_or_create_sheet(wb, "Entities",
                                    ["Entity Name", "Entity Code"])

    # ── Discover entity folders (auto-detects new ones) ──
    log.info("Scanning Drive folder structure...")
    subfolders = list_subfolders(drive, DRIVE_ROOT_ID)

    # Track results for summary email
    results = {
        "processed": [],    # (entity, filename, amount, confidence)
        "flagged": [],      # (entity, filename, confidence, notes)
        "skipped": [],      # (entity, filename, reason)
        "new_entities": []  # entity names auto-detected
    }

    # ── Process each entity folder ──
    log.info(f"All subfolders found: {list(subfolders.keys())}")
    for folder_name, folder_id in subfolders.items():

        # Skip system folders and known non-entity files/folders
        SKIP_NAMES = {"00_INBOX", "00_UNCATEGORIZED", "Filing Taxonomy", DONE_FOLDER}
        if folder_name.startswith("_") or folder_name.startswith("00_") or folder_name in SKIP_NAMES:
            log.info(f"Skipping system folder: {folder_name}")
            continue

        log.info(f"Processing folder: '{folder_name}'")
        
        # List files found
        files_found = list_files(drive, folder_id)
        log.info(f"  Files in '{folder_name}': {[f['name'] for f in files_found]}")

        # Auto-detect new entity
        if folder_name not in ENTITIES:
            log.info(f"New entity detected: {folder_name}")
            results["new_entities"].append(folder_name)
            ensure_entity_in_excel(ws_ents, folder_name)

        # Ensure /done subfolder exists
        done_id = ensure_subfolder(drive, folder_id, DONE_FOLDER)

        # List files in this entity folder
        files = files_found
        if not files:
            log.info(f"  No files in {folder_name}")
            continue

        for f in files:
            filename = f["name"]
            file_id  = f["id"]
            mime     = f.get("mimeType", "")

            log.info(f"  Reading: {filename} ({mime})")

            # Skip unsupported types
            if mime not in SUPPORTED_MIME:
                log.warning(f"  Skipping unsupported type: {mime}")
                results["skipped"].append((folder_name, filename, f"Unsupported type: {mime}"))
                continue

            # Download file
            try:
                file_bytes = download_file(drive, file_id).read()
            except Exception as e:
                log.error(f"  Download failed: {e}")
                results["skipped"].append((folder_name, filename, f"Download error: {e}"))
                continue

            # Filename-only duplicate check before AI call (saves API cost)
            dup, reason, dup_row = is_duplicate(ws_tx, filename)
            if dup:
                log.info(f"  → DUPLICATE (filename) — {reason}. Skipping.")
                results["skipped"].append((folder_name, filename, f"Duplicate: {reason}"))
                continue

            # AI extraction
            data = extract_invoice_data(file_bytes, filename, folder_name, mime)

            confidence = data.get("confidence", 0)
            vendor     = data.get("vendor", "")
            inv_num    = data.get("invoice_number", "N/A")
            is_paid    = data.get("is_paid", False)
            pay_date   = data.get("payment_date", None)
            groups     = data.get("line_groups", [])
            total_amt  = sum(g["amount"] for g in groups)

            log.info(f"  Confidence: {confidence*100:.0f}% | Vendor: {vendor} | "
                     f"Amount: ${total_amt} | Groups: {len(groups)} | Paid: {is_paid}")

            # ── Paid receipt — update existing rows, do not create new ones ──
            if is_paid:
                updated = 0
                for row_num in range(3, ws_tx.max_row + 1):
                    ex_inv    = str(ws_tx.cell(row=row_num, column=11).value or "").strip()
                    ex_vendor = str(ws_tx.cell(row=row_num, column=5).value or "").strip().lower()
                    if ex_inv == inv_num and (
                        ex_vendor == vendor.lower() or
                        ex_vendor in vendor.lower() or
                        vendor.lower() in ex_vendor
                    ):
                        update_status_to_paid(ws_tx, row_num, pay_date)
                        updated += 1
                if updated > 0:
                    log.info(f"  → PAID RECEIPT — updated {updated} row(s) to Paid")
                    results["skipped"].append((folder_name, filename, f"Paid receipt — updated {updated} row(s)"))
                    move_file(drive, file_id, done_id, folder_id)
                    continue
                # No matching rows found — fall through and process as new entry

            # ── Route based on confidence ──
            if confidence >= CONFIDENCE_THRESHOLD:
                rows_written = 0
                for group in groups:
                    group_data = {
                        **data,
                        "category":    group.get("category", "Other"),
                        "amount":      group["amount"],
                        "description": group.get("description", ""),
                    }
                    dup, reason, dup_row = is_duplicate(ws_tx, filename, data=group_data)
                    if dup:
                        log.info(f"  → Group duplicate ({group['category']}) — {reason}")
                        continue
                    append_transaction(ws_tx, folder_name, group_data, filename)
                    rows_written += 1
                    log.info(f"  → Written: {group['category']} ${group['amount']}")

                if rows_written > 0:
                    results["processed"].append((folder_name, filename, total_amt, confidence))
                    log.info(f"  → {rows_written} row(s) written to Transactions sheet")
                    # Update supplier learned patterns
                    try:
                        import json as _json
                        sups = {}
                        if os.path.exists("suppliers.json"):
                            with open("suppliers.json") as sf:
                                sups = _json.load(sf)
                        sup_key = f"{folder_name}|{vendor}"
                        profile = sups.get(sup_key, {})
                        learned = profile.get("_learned", [])
                        for g in groups:
                            entry = f"{g['category']}: {g.get('description','')}"
                            if entry not in learned:
                                learned.append(entry)
                        profile["_learned"] = learned
                        sups[sup_key] = profile
                        with open("suppliers.json", "w") as sf:
                            _json.dump(sups, sf, indent=2)
                    except Exception:
                        pass
                else:
                    results["skipped"].append((folder_name, filename, "All groups were duplicates"))
            else:
                append_review(ws_review, folder_name, data, filename)
                results["flagged"].append((
                    folder_name, filename, confidence,
                    data.get("notes", "Low confidence")
                ))
                log.info(f"  → Flagged for review (confidence {confidence*100:.0f}% < 90%)")

            # Move file to /done
            move_file(drive, file_id, done_id, folder_id)
            log.info(f"  → Moved to /done")

    # ── Save and upload Excel ──
    wb.save(excel_path)
    upload_excel(drive, excel_path)

    # ── Export portal data to GitHub ──
    export_portal_json(wb)

    # ── Send summary email ──
    send_summary_email(results)

    log.info("=" * 60)
    log.info("Run complete.")
    log.info("=" * 60)


# ─── Summary email ────────────────────────────────────────────────────────────
def send_summary_email(results):
    processed = results["processed"]
    flagged   = results["flagged"]
    skipped   = results["skipped"]
    new_ents  = results["new_entities"]

    total_amount = sum(r[2] for r in processed)
    today = datetime.today().strftime("%B %d, %Y")

    # Plain text
    lines = [
        f"Nagel Law — Nightly Expense Report",
        f"Run date: {today}",
        f"{'='*48}",
        f"",
        f"PROCESSED AUTOMATICALLY ({len(processed)} documents — ${total_amount:,.2f} total)",
    ]
    for entity, fname, amount, conf in processed:
        lines.append(f"  ✓  [{entity}]  {fname}  —  ${amount:,.2f}  ({conf*100:.0f}% confidence)")

    if flagged:
        lines += [
            f"",
            f"NEEDS YOUR REVIEW ({len(flagged)} documents — confidence below 90%)",
        ]
        for entity, fname, conf, notes in flagged:
            lines.append(f"  ⚠  [{entity}]  {fname}  —  {conf*100:.0f}%  |  {notes}")
        lines.append(f"  → Open the portal → 'Needs Review' tab to confirm or correct these.")

    if new_ents:
        lines += [
            f"",
            f"NEW ENTITIES DETECTED ({len(new_ents)})",
        ]
        for e in new_ents:
            lines.append(f"  +  {e}  — added to Excel and portal automatically")

    if skipped:
        lines += [
            f"",
            f"SKIPPED ({len(skipped)} files — unsupported format or error)",
        ]
        for entity, fname, reason in skipped:
            lines.append(f"  ✗  [{entity}]  {fname}  —  {reason}")

    lines += [
        f"",
        f"{'='*48}",
        f"Nagel Law Expense Portal — automated nightly run",
    ]

    body_text = "\n".join(lines)

    # HTML version
    def row_color(i): return "#f9f9f9" if i % 2 == 0 else "#ffffff"

    proc_rows = "".join(
        f'<tr style="background:{row_color(i)}">'
        f'<td style="padding:6px 10px">{entity}</td>'
        f'<td style="padding:6px 10px">{fname}</td>'
        f'<td style="padding:6px 10px;text-align:right">${amount:,.2f}</td>'
        f'<td style="padding:6px 10px;text-align:center;color:#15803d">{conf*100:.0f}%</td>'
        f'</tr>'
        for i, (entity, fname, amount, conf) in enumerate(processed)
    ) if processed else '<tr><td colspan="4" style="padding:10px;color:#aaa;text-align:center">No documents processed tonight</td></tr>'

    flag_rows = "".join(
        f'<tr style="background:{row_color(i)}">'
        f'<td style="padding:6px 10px">{entity}</td>'
        f'<td style="padding:6px 10px">{fname}</td>'
        f'<td style="padding:6px 10px;text-align:center;color:#b45309">{conf*100:.0f}%</td>'
        f'<td style="padding:6px 10px;color:#666">{notes}</td>'
        f'</tr>'
        for i, (entity, fname, conf, notes) in enumerate(flagged)
    ) if flagged else ""

    new_ent_section = ""
    if new_ents:
        items = "".join(f'<li style="margin:4px 0">{e} — added automatically</li>' for e in new_ents)
        new_ent_section = f"""
        <h3 style="color:#1a2f5a;margin:24px 0 8px">New Entities Detected</h3>
        <ul style="margin:0;padding-left:20px;color:#374151">{items}</ul>"""

    flagged_section = ""
    if flagged:
        flagged_section = f"""
        <h3 style="color:#b45309;margin:24px 0 8px">Needs Your Review ({len(flagged)})</h3>
        <p style="color:#666;font-size:13px;margin:0 0 10px">
          Confidence below 90% — open the portal's <strong>Needs Review</strong> tab to confirm or correct.
        </p>
        <table style="width:100%;border-collapse:collapse;font-size:13px">
          <thead>
            <tr style="background:#1a2f5a;color:#fff">
              <th style="padding:8px 10px;text-align:left">Entity</th>
              <th style="padding:8px 10px;text-align:left">File</th>
              <th style="padding:8px 10px;text-align:center">Confidence</th>
              <th style="padding:8px 10px;text-align:left">Notes</th>
            </tr>
          </thead>
          <tbody>{flag_rows}</tbody>
        </table>"""

    body_html = f"""
    <div style="font-family:'Helvetica Neue',Arial,sans-serif;max-width:640px;margin:0 auto;color:#1a1a1a">
      <div style="background:#1a2f5a;padding:24px 28px;border-radius:10px 10px 0 0">
        <div style="font-size:20px;font-weight:700;color:#fff">Nagel Law</div>
        <div style="font-size:13px;color:#c9a84c;letter-spacing:1px;text-transform:uppercase;margin-top:2px">
          Nightly Expense Report
        </div>
      </div>
      <div style="background:#f8f9fc;padding:18px 28px;border-left:1px solid #e4e7f0;border-right:1px solid #e4e7f0">
        <span style="font-size:13px;color:#888">{today}</span>
      </div>
      <div style="background:#fff;padding:24px 28px;border:1px solid #e4e7f0;border-top:none">

        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:14px;margin-bottom:24px">
          <div style="background:#f6f7fb;border-radius:8px;padding:14px 16px;border:1px solid #e4e7f0">
            <div style="font-size:10px;font-weight:700;color:#9ca3af;text-transform:uppercase;letter-spacing:.6px;margin-bottom:6px">Processed</div>
            <div style="font-size:24px;font-weight:700;color:#1a2f5a;font-family:monospace">{len(processed)}</div>
            <div style="font-size:11px;color:#9ca3af;margin-top:3px">documents tonight</div>
          </div>
          <div style="background:#f6f7fb;border-radius:8px;padding:14px 16px;border:1px solid #e4e7f0">
            <div style="font-size:10px;font-weight:700;color:#9ca3af;text-transform:uppercase;letter-spacing:.6px;margin-bottom:6px">Total Amount</div>
            <div style="font-size:24px;font-weight:700;color:#1a2f5a;font-family:monospace">${total_amount:,.0f}</div>
            <div style="font-size:11px;color:#9ca3af;margin-top:3px">auto-added to Excel</div>
          </div>
          <div style="background:{'#fef3c7' if flagged else '#f6f7fb'};border-radius:8px;padding:14px 16px;border:1px solid {'#fcd34d' if flagged else '#e4e7f0'}">
            <div style="font-size:10px;font-weight:700;color:#9ca3af;text-transform:uppercase;letter-spacing:.6px;margin-bottom:6px">Needs Review</div>
            <div style="font-size:24px;font-weight:700;color:{'#b45309' if flagged else '#1a2f5a'};font-family:monospace">{len(flagged)}</div>
            <div style="font-size:11px;color:#9ca3af;margin-top:3px">below 90% confidence</div>
          </div>
        </div>

        <h3 style="color:#1a2f5a;margin:0 0 10px">Processed Tonight ({len(processed)})</h3>
        <table style="width:100%;border-collapse:collapse;font-size:13px">
          <thead>
            <tr style="background:#1a2f5a;color:#fff">
              <th style="padding:8px 10px;text-align:left">Entity</th>
              <th style="padding:8px 10px;text-align:left">File</th>
              <th style="padding:8px 10px;text-align:right">Amount</th>
              <th style="padding:8px 10px;text-align:center">Confidence</th>
            </tr>
          </thead>
          <tbody>{proc_rows}</tbody>
        </table>

        {flagged_section}
        {new_ent_section}

      </div>
      <div style="background:#f6f7fb;padding:14px 28px;border-radius:0 0 10px 10px;border:1px solid #e4e7f0;border-top:none">
        <span style="font-size:11px;color:#bbb">Nagel Law Expense Portal — automated nightly run · andres@nagellaw.com</span>
      </div>
    </div>"""

    # Send
    smtp_user = os.environ.get("SMTP_USER", "")
    smtp_pass = os.environ.get("SMTP_PASSWORD", "")

    if not smtp_user or not smtp_pass:
        log.warning("SMTP credentials not set — skipping email.")
        log.info("Summary:\n" + body_text)
        return

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"Nagel Law — Expense Report {today} ({len(processed)} docs · ${total_amount:,.0f})"
    msg["From"]    = smtp_user
    msg["To"]      = SUMMARY_EMAIL
    msg.attach(MIMEText(body_text, "plain"))
    msg.attach(MIMEText(body_html, "html"))

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, SUMMARY_EMAIL, msg.as_string())
        log.info(f"Summary email sent to {SUMMARY_EMAIL}")
    except Exception as e:
        log.error(f"Email send failed: {e}")


if __name__ == "__main__":
    run()
